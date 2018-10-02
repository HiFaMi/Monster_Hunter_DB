import openpyxl

wb = openpyxl.load_workbook('armor_list.xlsx')
ws = wb['Skill']


for i in range(1, 144):
    base = ws.cell(row=i, column=2).value
    try:
        skill_list = base.split('|')

    except AttributeError:
        ws.cell(row=i, column=3).value = base

    else:
        for index in range(len(skill_list)):
            ws.cell(row=i, column=3+index).value = skill_list[index]

wb.save('armor_list.xlsx')
print('Rework Done')

