import os

import django
import openpyxl

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from weapon.models import *

WB = openpyxl.load_workbook('weapon_data.xlsx')


def great_sword():

    ws = WB['Great_Sword']

    for row in range(2, 92):
            name = ws.cell(row=row, column=1).value
            rare = ws.cell(row=row, column=2).value
            produce = ws.cell(row=row, column=3).value
            attack = ws.cell(row=row, column=4).value
            defense = ws.cell(row=row, column=5).value
            if ws.cell(row=row, column=6).value is None:
                weapon_property = '없음'
                weapon_property_fiqure = '없음'
            else:
                weapon_property = ws.cell(row=row, column=6).value
                weapon_property_fiqure = ws.cell(row=row, column=7).value

            if ws.cell(row=row, column=8).value is None:
                abnormal_condition = '없음'
                abnormal_condition_figure = '없음'
            else:
                abnormal_condition = ws.cell(row=row, column=8).value
                abnormal_condition_figure = ws.cell(row=row, column=9).value

            fatal_blow = ws.cell(row=row, column=10).value
            if ws.cell(row=row, column=11).value is None:
                slot = '없음'
            else:
                slot = ws.cell(row=row, column=11).value

            GreatSword.objects.create(
                name=name,
                rare=rare,
                produce=produce,
                attack=attack,
                defense=defense,
                weapon_property=weapon_property,
                weapon_property_fiqure=weapon_property_fiqure,
                abnormal_condition=abnormal_condition,
                abnormal_condition_figure=abnormal_condition_figure,
                fatal_blow=fatal_blow,
                slot=slot,
                # cost=cost
            )
    print('Great Sword Done')


def long_sword():

    ws = WB['Long_Sword']

    for row in range(1, 81):
        name = ws.cell(row=row, column=1).value
        rare = ws.cell(row=row, column=2).value
        produce = ws.cell(row=row, column=3).value
        attack = ws.cell(row=row, column=4).value
        defense = ws.cell(row=row, column=5).value
        if ws.cell(row=row, column=6).value is None:
            weapon_property = '없음'
            weapon_property_fiqure = '없음'
        else:
            weapon_property = ws.cell(row=row, column=6).value
            weapon_property_fiqure = ws.cell(row=row, column=7).value

        if ws.cell(row=row, column=8).value is None:
            abnormal_condition = '없음'
            abnormal_condition_figure = '없음'
        else:
            abnormal_condition = ws.cell(row=row, column=8).value
            abnormal_condition_figure = ws.cell(row=row, column=9).value

        fatal_blow = ws.cell(row=row, column=10).value
        if ws.cell(row=row, column=11).value is None:
            slot = '없음'
        else:
            slot = ws.cell(row=row, column=11).value

        LongSword.objects.create(
            name=name,
            rare=rare,
            produce=produce,
            attack=attack,
            defense=defense,
            weapon_property=weapon_property,
            weapon_property_fiqure=weapon_property_fiqure,
            abnormal_condition=abnormal_condition,
            abnormal_condition_figure=abnormal_condition_figure,
            fatal_blow=fatal_blow,
            slot=slot,
        )
    print('Long Sword Done')


def one_hand_sword():

    ws = WB['One_Hand_Sword']

    for row in range(1, 93):
        name = ws.cell(row=row, column=1).value
        rare = ws.cell(row=row, column=2).value
        produce = ws.cell(row=row, column=3).value
        attack = ws.cell(row=row, column=4).value
        defense = ws.cell(row=row, column=5).value
        if ws.cell(row=row, column=6).value is None:
            weapon_property = '없음'
            weapon_property_fiqure = '없음'
        else:
            weapon_property = ws.cell(row=row, column=6).value
            weapon_property_fiqure = ws.cell(row=row, column=7).value

        if ws.cell(row=row, column=8).value is None:
            abnormal_condition = '없음'
            abnormal_condition_figure = '없음'
        else:
            abnormal_condition = ws.cell(row=row, column=8).value
            abnormal_condition_figure = ws.cell(row=row, column=9).value

        fatal_blow = ws.cell(row=row, column=10).value
        if ws.cell(row=row, column=11).value is None:
            slot = '없음'
        else:
            slot = ws.cell(row=row, column=11).value

        OneHandSword.objects.create(
            name=name,
            rare=rare,
            produce=produce,
            attack=attack,
            defense=defense,
            weapon_property=weapon_property,
            weapon_property_fiqure=weapon_property_fiqure,
            abnormal_condition=abnormal_condition,
            abnormal_condition_figure=abnormal_condition_figure,
            fatal_blow=fatal_blow,
            slot=slot,
        )
    print('One Hand Sword Done')


def two_swords():

    ws = WB['Two_Swords']

    for row in range(1, 84):
        name = ws.cell(row=row, column=1).value
        rare = ws.cell(row=row, column=2).value
        produce = ws.cell(row=row, column=3).value
        attack = ws.cell(row=row, column=4).value
        defense = ws.cell(row=row, column=5).value
        if ws.cell(row=row, column=6).value is None:
            weapon_property = '없음'
            weapon_property_fiqure = '없음'
        else:
            weapon_property = ws.cell(row=row, column=6).value
            weapon_property_fiqure = ws.cell(row=row, column=7).value

        if ws.cell(row=row, column=8).value is None:
            abnormal_condition = '없음'
            abnormal_condition_figure = '없음'
        else:
            abnormal_condition = ws.cell(row=row, column=8).value
            abnormal_condition_figure = ws.cell(row=row, column=9).value

        fatal_blow = ws.cell(row=row, column=10).value
        if ws.cell(row=row, column=11).value is None:
            slot = '없음'
        else:
            slot = ws.cell(row=row, column=11).value

        TwoSwords.objects.create(
            name=name,
            rare=rare,
            produce=produce,
            attack=attack,
            defense=defense,
            weapon_property=weapon_property,
            weapon_property_fiqure=weapon_property_fiqure,
            abnormal_condition=abnormal_condition,
            abnormal_condition_figure=abnormal_condition_figure,
            fatal_blow=fatal_blow,
            slot=slot,
        )
    print('Two Swords Done')


def hammer():

    ws = WB['Hammer']

    for row in range(1, 85):
        name = ws.cell(row=row, column=1).value
        rare = ws.cell(row=row, column=2).value
        produce = ws.cell(row=row, column=3).value
        attack = ws.cell(row=row, column=4).value
        defense = ws.cell(row=row, column=5).value
        if ws.cell(row=row, column=6).value is None:
            weapon_property = '없음'
            weapon_property_fiqure = '없음'
        else:
            weapon_property = ws.cell(row=row, column=6).value
            weapon_property_fiqure = ws.cell(row=row, column=7).value

        if ws.cell(row=row, column=8).value is None:
            abnormal_condition = '없음'
            abnormal_condition_figure = '없음'
        else:
            abnormal_condition = ws.cell(row=row, column=8).value
            abnormal_condition_figure = ws.cell(row=row, column=9).value

        fatal_blow = ws.cell(row=row, column=10).value
        if ws.cell(row=row, column=11).value is None:
            slot = '없음'
        else:
            slot = ws.cell(row=row, column=11).value

        Hammer.objects.create(
            name=name,
            rare=rare,
            produce=produce,
            attack=attack,
            defense=defense,
            weapon_property=weapon_property,
            weapon_property_fiqure=weapon_property_fiqure,
            abnormal_condition=abnormal_condition,
            abnormal_condition_figure=abnormal_condition_figure,
            fatal_blow=fatal_blow,
            slot=slot,
        )
    print('Hammer Done')


def lance():

    ws = WB['Lance']

    for row in range(1, 85):
        name = ws.cell(row=row, column=1).value
        rare = ws.cell(row=row, column=2).value
        produce = ws.cell(row=row, column=3).value
        attack = ws.cell(row=row, column=4).value
        defense = ws.cell(row=row, column=5).value
        if ws.cell(row=row, column=6).value is None:
            weapon_property = '없음'
            weapon_property_fiqure = '없음'
        else:
            weapon_property = ws.cell(row=row, column=6).value
            weapon_property_fiqure = ws.cell(row=row, column=7).value

        if ws.cell(row=row, column=8).value is None:
            abnormal_condition = '없음'
            abnormal_condition_figure = '없음'
        else:
            abnormal_condition = ws.cell(row=row, column=8).value
            abnormal_condition_figure = ws.cell(row=row, column=9).value

        fatal_blow = ws.cell(row=row, column=10).value
        if ws.cell(row=row, column=11).value is None:
            slot = '없음'
        else:
            slot = ws.cell(row=row, column=11).value

        Lance.objects.create(
            name=name,
            rare=rare,
            produce=produce,
            attack=attack,
            defense=defense,
            weapon_property=weapon_property,
            weapon_property_fiqure=weapon_property_fiqure,
            abnormal_condition=abnormal_condition,
            abnormal_condition_figure=abnormal_condition_figure,
            fatal_blow=fatal_blow,
            slot=slot,
        )
    print('Lance Done')


def gun_lance():

    ws = WB['Gun_Lance']

    for row in range(1, 80):
        name = ws.cell(row=row, column=1).value
        rare = ws.cell(row=row, column=2).value
        produce = ws.cell(row=row, column=3).value
        attack = ws.cell(row=row, column=4).value
        defense = ws.cell(row=row, column=5).value
        if ws.cell(row=row, column=6).value is None:
            weapon_property = '없음'
            weapon_property_fiqure = '없음'
        else:
            weapon_property = ws.cell(row=row, column=6).value
            weapon_property_fiqure = ws.cell(row=row, column=7).value

        if ws.cell(row=row, column=8).value is None:
            abnormal_condition = '없음'
            abnormal_condition_figure = '없음'
        else:
            abnormal_condition = ws.cell(row=row, column=8).value
            abnormal_condition_figure = ws.cell(row=row, column=9).value

        fatal_blow = ws.cell(row=row, column=10).value
        if ws.cell(row=row, column=11).value is None:
            slot = '없음'
        else:
            slot = ws.cell(row=row, column=11).value

        cannon = ws.cell(row, column=12).value

        GunLance.objects.create(
            name=name,
            rare=rare,
            produce=produce,
            attack=attack,
            defense=defense,
            weapon_property=weapon_property,
            weapon_property_fiqure=weapon_property_fiqure,
            abnormal_condition=abnormal_condition,
            abnormal_condition_figure=abnormal_condition_figure,
            fatal_blow=fatal_blow,
            slot=slot,
            cannon=cannon,
        )
    print('Gun Lance Done')


def switch_axe():

    ws = WB['Switch_Axe']

    for row in range(1, 80):
        name = ws.cell(row=row, column=1).value
        rare = ws.cell(row=row, column=2).value
        produce = ws.cell(row=row, column=3).value
        attack = ws.cell(row=row, column=4).value
        defense = ws.cell(row=row, column=5).value
        if ws.cell(row=row, column=6).value is None:
            weapon_property = '없음'
            weapon_property_fiqure = '없음'
        else:
            weapon_property = ws.cell(row=row, column=6).value
            weapon_property_fiqure = ws.cell(row=row, column=7).value

        if ws.cell(row=row, column=8).value is None:
            abnormal_condition = '없음'
            abnormal_condition_figure = '없음'
        else:
            abnormal_condition = ws.cell(row=row, column=8).value
            abnormal_condition_figure = ws.cell(row=row, column=9).value

        fatal_blow = ws.cell(row=row, column=10).value
        if ws.cell(row=row, column=11).value is None:
            slot = '없음'
        else:
            slot = ws.cell(row=row, column=11).value

        phial_type = ws.cell(row, column=12).value

        SwitchAxe.objects.create(
            name=name,
            rare=rare,
            produce=produce,
            attack=attack,
            defense=defense,
            weapon_property=weapon_property,
            weapon_property_fiqure=weapon_property_fiqure,
            abnormal_condition=abnormal_condition,
            abnormal_condition_figure=abnormal_condition_figure,
            fatal_blow=fatal_blow,
            slot=slot,
            phial_type=phial_type,
        )
    print('Switch Axe Done')


def charge_blade():

    ws = WB['Charge_Blade']

    for row in range(1, 76):
        name = ws.cell(row=row, column=1).value
        rare = ws.cell(row=row, column=2).value
        produce = ws.cell(row=row, column=3).value
        attack = ws.cell(row=row, column=4).value
        defense = ws.cell(row=row, column=5).value
        if ws.cell(row=row, column=6).value is None:
            weapon_property = '없음'
            weapon_property_fiqure = '없음'
        else:
            weapon_property = ws.cell(row=row, column=6).value
            weapon_property_fiqure = ws.cell(row=row, column=7).value

        if ws.cell(row=row, column=8).value is None:
            abnormal_condition = '없음'
            abnormal_condition_figure = '없음'
        else:
            abnormal_condition = ws.cell(row=row, column=8).value
            abnormal_condition_figure = ws.cell(row=row, column=9).value

        fatal_blow = ws.cell(row=row, column=10).value
        if ws.cell(row=row, column=11).value is None:
            slot = '없음'
        else:
            slot = ws.cell(row=row, column=11).value

        phial_type = ws.cell(row, column=12).value

        ChargeBlade.objects.create(
            name=name,
            rare=rare,
            produce=produce,
            attack=attack,
            defense=defense,
            weapon_property=weapon_property,
            weapon_property_fiqure=weapon_property_fiqure,
            abnormal_condition=abnormal_condition,
            abnormal_condition_figure=abnormal_condition_figure,
            fatal_blow=fatal_blow,
            slot=slot,
            phial_type=phial_type,
        )
    print('Charge Blade Done')


def light_bowgun():

    ws = WB['Light_Bowgun']

    for row in range(1, 85):
        name = ws.cell(row=row, column=1).value
        rare = ws.cell(row=row, column=2).value
        produce = ws.cell(row=row, column=3).value
        attack = ws.cell(row=row, column=4).value
        recoil = ws.cell(row=row, column=5).value
        fatal_blow = ws.cell(row=row, column=6).value
        if ws.cell(row=row, column=7).value is None:
            slot = '없음'
        else:
            slot = ws.cell(row=row, column=7).value
        custom_mod = ws.cell(row=row, column=8).value
        special_ammo = ws.cell(row, column=9).value

        LightBowgun.objects.create(
            name=name,
            rare=rare,
            produce=produce,
            attack=attack,
            recoil=recoil,
            fatal_blow=fatal_blow,
            slot=slot,
            custom_mod=custom_mod,
            special_ammo=special_ammo,
        )
    print('Light Bowgun Done')


if __name__ == "__main__":

    great_sword()
    long_sword()
    one_hand_sword()
    two_swords()
    hammer()
    lance()
    gun_lance()
    switch_axe()
    light_bowgun()
