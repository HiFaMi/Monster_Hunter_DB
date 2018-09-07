import requests
from bs4 import BeautifulSoup
import openpyxl

url = 'http://mhf.inven.co.kr/dataninfo/skill/mhw.php'

skill_url = requests.get(url)
html = skill_url.text

soup = BeautifulSoup(html, 'lxml')
skill_name_list = soup.select('tbody > tr > td.name > b')
skill_list = soup.select('tbody > tr > td.etc0')

wb = openpyxl.load_workbook('armor_list.xlsx')
ws = wb['Skill']

for index in range(len(skill_name_list)):
    ws.cell(row=index+1, column=1).value = skill_name_list[index].get_text(strip=True)
    ws.cell(row=index+1, column=2).value = skill_list[index].get_text(strip=True)
    print(skill_name_list[index].get_text(strip=True) + ' Done')

wb.save('armor_list.xlsx')
print('all saved')
