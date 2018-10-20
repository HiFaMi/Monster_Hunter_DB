import openpyxl
import os
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from armor.models.armor import Armor
from items.models.accessory import Accessory
from items.models import Items
from skill.models import Skill

wb = openpyxl.load_workbook('armor_list.xlsx')


def armor_model():
    ws = wb['Armor']

    for index in range(1, 754):
        name = ws.cell(row=index, column=1).value
        equip_region = ws.cell(row=index, column=2).value
        rare = ws.cell(row=index, column=3).value
        cost = ws.cell(row=index, column=4).value
        defense = ws.cell(row=index, column=5).value
        vs_fire = ws.cell(row=index, column=7).value
        vs_water = ws.cell(row=index, column=8).value
        vs_thunder = ws.cell(row=index, column=9).value
        vs_ice = ws.cell(row=index, column=10).value
        vs_dragon = ws.cell(row=index, column=11).value
        slot = ws.cell(row=index, column=12).value

        Armor.objects.create(
            name=name,
            equip_region=equip_region,
            rare=rare,
            cost=cost,
            defense=defense,
            vs_fire=vs_fire,
            vs_water=vs_water,
            vs_thunder=vs_thunder,
            vs_ice=vs_ice,
            vs_dragon=vs_dragon,
            slot=slot,
        )

    print('Armor Done')


def accessory_model():
    ws = wb['Accessory']

    for index in range(1, 99):
        name = ws.cell(row=index, column=1). value
        slot_lv = ws.cell(row=index, column=2). value
        rare = ws.cell(row=index, column=3). value

        Accessory.objects.create(
            name=name,
            slot_lv=slot_lv,
            rare=rare,
        )
    print('Accessory Done')


def items_model():
    ws = wb['Item']

    for index in range(1, 393):
        name = ws.cell(row=index, column=1).value

        Items.objects.create(
            name=name,
        )

    print('Items Done')


def skill_model():
    ws = wb['Skill']

    for row in range(1, 144):
        name = ws.cell(row=row, column=1).value
        Lv1 = ws.cell(row=row, column=3).value
        Lv2 = ws.cell(row=row, column=4).value
        Lv3 = ws.cell(row=row, column=5).value
        Lv4 = ws.cell(row=row, column=6).value
        Lv5 = ws.cell(row=row, column=7).value

        Skill.objects.create(
            name=name,
            Lv1=Lv1,
            Lv2=Lv2,
            Lv3=Lv3,
            Lv4=Lv4,
            Lv5=Lv5,
        )

    print('Skill Done')


if __name__ == '__main__':
    armor_model()
    accessory_model()
    items_model()
    skill_model()
