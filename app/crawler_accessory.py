import requests
from bs4 import BeautifulSoup
import openpyxl

url = 'http://mhf.inven.co.kr/dataninfo/mhw/accessory/'

accessory_url = requests.get(url)
html = accessory_url.text

soup = BeautifulSoup(html, 'lxml')

accessory_name_list = soup.select('td.name.left > b')
accessory_slot_level_list = soup.select('td.slot')
accessory_slot_rare_list = soup.select('td.rare')
accessory_skill_list = soup.select('td.skill.text_left > b')

wb = openpyxl.load_workbook('armor_list.xlsx')
ws = wb['Accessory']

for index in range(len(accessory_name_list)):
    ws.cell(row=index+1, column=1).value = accessory_name_list[index].get_text(strip=True)
    ws.cell(row=index+1, column=2).value = accessory_slot_level_list[index].get_text(strip=True)
    ws.cell(row=index + 1, column=3).value = accessory_slot_rare_list[index].get_text(strip=True)
    ws.cell(row=index + 1, column=4).value = accessory_skill_list[index].get_text(strip=True)
    print(accessory_name_list[index].get_text(strip=True) + ' Done')

wb.save('armor_list.xlsx')
print('all saved')


