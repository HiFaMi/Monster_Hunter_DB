import openpyxl

wb = openpyxl.load_workbook('armor_list.xlsx')
ws = wb['Armor']


def replace_multiple(main_string, to_be_replaces, new_string):
    for elem in to_be_replaces:
        if elem in main_string:
            main_string = main_string.replace(elem, new_string)
    return main_string


for i in range(1, 144):
    base = ws.cell(row=i, column=6).value
    replace_base = replace_multiple(base, ['화', '수', '뇌', '빙', '용'], '')
    armor_list = replace_base.split(':')
    for list_index in range(1, len(armor_list)):
        ws.cell(row=i, column=6+list_index).value = armor_list[list_index]


wb.save('armor_list.xlsx')
print('Rework Done')
