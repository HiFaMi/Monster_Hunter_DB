import openpyxl

wb = openpyxl.load_workbook('weapon_data.xlsx')

ws = wb['Great_Sword']


for row in range(2, 92):
    for col in range(1, len(ws['A:U'])):
        print(ws.cell(row=row, column=col).value)


